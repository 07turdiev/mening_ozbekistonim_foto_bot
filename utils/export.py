"""Arizalarni Excel (.xlsx) va ZIP ko'rinishida eksport qilish."""
from __future__ import annotations

import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import cfg
from utils.validators import safe_name

STATUS_UZ = {"pending": "Ko‘rib chiqilmoqda", "approved": "Qabul qilindi", "rejected": "Rad etildi"}

HEADERS = [
    ("№", 6),
    ("ID", 7),
    ("F.I.Sh.", 30),
    ("Telefon", 16),
    ("Telegram", 16),
    ("Fotosurat nomi", 28),
    ("Suratga olingan joy", 30),
    ("Sana", 12),
    ("Izoh", 55),
    ("O‘lcham (px)", 14),
    ("Hajm (MB)", 11),
    ("EXIF", 10),
    ("Kamera ma’lumoti", 30),
    ("Holat", 18),
    ("Rad etish sababi", 28),
    ("Fayl nomi", 34),
    ("Yuborilgan vaqt", 19),
]

HEAD_FILL = PatternFill("solid", fgColor="1F4E79")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
STATUS_FILL = {
    "approved": PatternFill("solid", fgColor="E2EFDA"),
    "rejected": PatternFill("solid", fgColor="FCE4E4"),
    "pending": PatternFill("solid", fgColor="FFF2CC"),
}


def build_excel(rows, status_filter: str | None = None) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Arizalar"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    title = ws.cell(row=1, column=1, value="“MENING O‘ZBEKISTONIM” FOTOTANLOVI — ishtirokchilar ro‘yxati")
    title.font = Font(bold=True, size=14, color="1F4E79")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    sub = ws.cell(row=2, column=1, value=f"Shakllantirildi: {cfg.now():%d.%m.%Y %H:%M} | Jami yozuvlar: {len(rows)}")
    sub.alignment = Alignment(horizontal="center")

    for col, (name, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col, value=name)
        cell.fill, cell.font, cell.border = HEAD_FILL, HEAD_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"

    for idx, r in enumerate(rows, start=1):
        size_mb = round((r["file_size"] or 0) / 1048576, 2)
        username = f"@{r['username']}" if r["username"] else str(r["user_id"])
        values = [
            idx, r["id"], r["fio"], r["phone"], username, r["title"], r["place"],
            r["shot_date"], r["description"], f"{r['width']}×{r['height']}", size_mb,
            "bor" if r["has_exif"] else "yo‘q", r["exif_info"] or "",
            STATUS_UZ.get(r["status"], r["status"]), r["reject_reason"] or "",
            r["file_name"] or "", r["created_at"],
        ]
        row_no = idx + 3
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_no, column=col, value=value)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=col in (3, 6, 7, 9, 13, 15, 16))
            fill = STATUS_FILL.get(r["status"])
            if fill and col == 14:
                cell.fill = fill

    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}{len(rows) + 3}"

    suffix = f"_{status_filter}" if status_filter else ""
    path = cfg.exports_dir / f"arizalar{suffix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(path)
    return path


def build_zip(rows, max_bytes: int = 45 * 1024 * 1024) -> tuple[Path, int, int]:
    """Suratlarni ZIP arxivga yig'adi. (arxiv, qo'shilgan, tashlab ketilgan) qaytaradi."""
    path = cfg.exports_dir / f"fotosuratlar_{datetime.now():%Y%m%d_%H%M%S}.zip"
    added = skipped = 0
    total = 0
    with zipfile.ZipFile(path, "w", zipfile.ZIP_STORED) as zf:
        for r in rows:
            src = Path(r["file_path"] or "")
            if not src.exists():
                skipped += 1
                continue
            size = src.stat().st_size
            if total + size > max_bytes:
                skipped += 1
                continue
            arc = f"{r['id']:04d}_{safe_name(r['fio'])}_{safe_name(r['title'])}{src.suffix}"
            zf.write(src, arc)
            total += size
            added += 1
    return path, added, skipped
